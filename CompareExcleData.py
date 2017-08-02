import openpyxl
from openpyxl.styles import PatternFill

def ReadExcelData(file_name):
    d = {}  # id to name map
    d1 = {}  # name to id map
    wb = openpyxl.load_workbook(file_name, data_only=True) # data_only flag will return the value not the formula
    for ws in wb.worksheets:
        for x in range(2, ws.max_row + 1):  # start from 2 to skip the title
            name = ws.cell(row=x, column=1).value
            id = ws.cell(row=x, column=2).value

            if (id in d):
                print ('duplicate id', id)
            else:
                d[id] = name

            # if (name in d1):
            #     print ('duplicate name', name)

            # name to id map 
            d1.setdefault(name,[]).append(id)

    # for k, v in d.items():
    #     print(k, v)
    print ('Total id count:', len(d))
    print ('Total name count:', len(d1))

    return d, d1

def main():
    print('Reading Data...')
    idToName, nameToId = ReadExcelData('公安.xlsx')

    print('Checking Data...')
    wb = openpyxl.load_workbook('发放.xlsx', data_only=True) # data_only flag will return the value not the formula

    correct = wrong = nonexist = 0; # counting

    for ws in wb.worksheets:
        for x in range(2, ws.max_row + 1):  # start from 2 to skip the title
            name = ws.cell(row=x, column=1).value
            id = ws.cell(row=x, column=2).value

            if (id in idToName):  # id exists
                if name != idToName[id]: # but name is not the same
                    print('身份证', id, '对应的姓名不正确。', name, '应该是', idToName[id])
                    wrong += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=3).value = '姓名不正确'
                    ws.cell(row=x, column=4).value = idToName[id]
                else:
                    correct += 1
            else: # cannot find id, it could be id number wrong, or id does not exist.
                if (name in nameToId):
                    print('姓名存在，身份证号码不存在:', name)
                    wrong += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

                    ws.cell(row=x, column=3).value = '身份证号码不正确'
                    for i, v in enumerate(nameToId[name]):
                        ws.cell(row=x, column=4+i).value = v
                        print('身份证号码可能是:', v)
                else:
                    print('姓名和身份证号码都不存在:', name, id)
                    nonexist += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    ws.cell(row=x, column=3).value = '姓名和身份证号码都不存在'

    print('Correct:', correct)
    print('Wrong:', wrong)
    print('Non-Exist:', nonexist)

    wb.save('new_document.xlsx')

if __name__ == '__main__':
    main()



