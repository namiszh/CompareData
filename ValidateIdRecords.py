#!/usr/bin/python

import openpyxl
from openpyxl.styles import PatternFill
import click
import datetime
import os.path

def ReadExcelData(file_name):
    print('Reading Data from', file_name, '...')
    d = {}  # id to name map
    d1 = {}  # name to id multiple map, 
    wb = openpyxl.load_workbook(file_name, data_only=True) # data_only flag will return the value not the formula
    for ws in wb.worksheets:
        for x in range(2, ws.max_row + 1):  # start from 2 to skip the title
            name = ws.cell(row=x, column=1).value
            id = ws.cell(row=x, column=2).value

            if (id in d):
                print ('ALERT: duplicate id:', id)
            else:
                d[id] = name

            # name to id map 
            d1.setdefault(name,[]).append(id)

    # for k, v in d.items():
    #     print(k, v)
    print ('Total id count:', len(d))
    print ('Total name count:', len(d1))

    return d, d1


@click.command()
@click.option('--database', prompt='The excel file used as database', help='The excel file used as database')
@click.option('--file', prompt='The excel file to check', help='The excel file to check')
def main(database, file):
    """Simple python program that Checks whether data records from one excel file are correct or not."""
    idToName, nameToId = ReadExcelData(database)

    print('Checking Data for file', file ,'...')
    wb = openpyxl.load_workbook(file, data_only=True) # data_only flag will return the value not the formula

    correct = wrong = nonexist = 0; # counting

    for ws in wb.worksheets:
        for x in range(2, ws.max_row + 1):  # start from 2 to skip the title
            name = ws.cell(row=x, column=1).value
            id = ws.cell(row=x, column=2).value

            if (id in idToName):  # id exists
                if name != idToName[id]: # but name is not the same
                    wrong += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=3).value = '姓名不正确'
                    ws.cell(row=x, column=4).value = idToName[id]
                    print('身份证', id, '对应的姓名不正确。', name, '应该是', idToName[id])
                else:
                    correct += 1
            else: # cannot find id, it could be id number wrong, or id does not exist.
                if (name in nameToId):
                    wrong += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
                    ws.cell(row=x, column=3).value = '身份证号码不正确'
                    print('身份证号码',id,'不存在. 查到名为',name,'的如下身份证号码:')
                    for i, v in enumerate(nameToId[name]):
                        ws.cell(row=x, column=4+i).value = v
                        print(v)
                else:
                    nonexist += 1
                    ws.cell(row=x, column=1).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    ws.cell(row=x, column=2).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    ws.cell(row=x, column=3).value = '姓名和身份证号码都不存在'
                    print('查无此人:', name, id)

    print('Correct:', correct)
    print('Wrong:', wrong)
    print('Non-Exist:', nonexist)

    # output to a new file
    dirName, baseName = os.path.split(file)
    shortName = os.path.splitext(baseName)[0].strip()
    extName = os.path.splitext(baseName)[1][1:].strip()
    now = datetime.datetime.now().strftime("%Y%m%d%H%M")
    newBaseName = shortName+now+'.'+extName
    newFileName = os.path.join(os.sep, dirName, newBaseName)
    print('Save result to', newFileName)
    wb.save(newFileName)

if __name__ == '__main__':
    main()



